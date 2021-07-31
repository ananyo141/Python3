#! python3
# This script updates the prices for garlic, celery or lemons in the excel file 'produceSales.xlsx'.
import openpyxl
from openpyxl.styles import Font

updatedPricing = {'Garlic': 3.07,
                  'Celery': 1.19,
                  'Lemon' : 1.27}

def main():
    print("Opening file")
    xlFile = openpyxl.load_workbook('./produceSales.xlsx')
    sheet = xlFile.active
    print("Searching and updating entries...")
    boldFont = Font(name = 'Times New Roman', bold = True)
    italicFont = Font(name = 'Verdana', size = 10.5, italic = True)
    for row in range(2, sheet.max_row + 1):
        produceCell = sheet.cell(row = row, column = 1)
        costCell = sheet.cell(row = row, column = 2)
        if produceCell.value in updatedPricing:
            costCell.value = updatedPricing[produceCell.value]
            produceCell.font = boldFont
            costCell.font = italicFont
    
    xlFile.save('./updatedProduceSales.xlsx')
    print("File updated and saved")


if __name__ == '__main__':
    main()