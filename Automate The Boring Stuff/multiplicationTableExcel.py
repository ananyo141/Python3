# Take a number from the command line and create NxN multiplication table in xlsx format
import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

CELL_SIZE = 8.43

def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: Include the number in commandline to save its multipicaton table")
    try:
        number = int(sys.argv[1])
    except:
        sys.exit("Integer expected")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # creating title
    sheet.title = 'Multiplication Table'
    sheet.merge_cells('A1:' + get_column_letter(number + 1) + '1')
    sheet['A1'] = 'Multiplication Table'.center(int(CELL_SIZE * number))
    sheet['A1'].font = Font(name = 'Courier New', bold = True)   # monospace font required

    sheet['A2'] = 'N = ' + str(number)
    sheet['A2'].font = Font(italic = True)

    # create the multiplication table number-header
    bold_italic_font = Font(bold = True, italic = True)
    for i in range(1, number + 1):

        # set row value
        row_cell = sheet.cell(row = 2, column = i + 1)
        row_cell.value = i
        row_cell.font = bold_italic_font

        # set column value
        column_cell = sheet.cell(row = i + 2, column = 1)
        column_cell.value = i
        column_cell.font = bold_italic_font

        # make the cells square
        sheet.row_dimensions[i + 2].height = 30

    # fill in the table
    for row in range(1, number + 1):
        for column in range(1, number + 1):
            sheet.cell(row = row + 2, column = column + 1).value = row * column
    
    workbook.save('./multiplicationTable.xlsx')


if __name__ == '__main__':
    main()